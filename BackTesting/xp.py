import os
from openpyxl import Workbook, load_workbook
import PyPDF2
import re
import pandas as pd

# Define the folder where the PDF files are located
pdf_folder = r'C:\Python\pythonProject\automacaob3\BackTesting\notas'

# Create a new Excel workbook
wb = Workbook()
sheet = wb.active

# Add headers to the columns
sheet['A1'] = 'Data' # type: ignore
sheet['B1'] = 'Valor Total' # type: ignore
sheet['C1'] = 'Credito-Debito' # type: ignore



# Iterate over all the files in the folder
for filename in os.listdir(pdf_folder):
    if filename.endswith('.pdf'):
        file_path = os.path.join(pdf_folder, filename)

        # Open the PDF file using PdfReader
        with open(file_path, 'rb') as pdf_file:
            # print(pdf_file.read())
            reader = PyPDF2.PdfReader(pdf_file)
            for page in reader.pages:
                # Extraia o texto da página atual
                page_text = page.extract_text()
                print(page_text)
                page_text2 = page_text.replace(' |','').split('\n')
                # Use expressões regulares para encontrar a data e o texto subsequente
                match = re.search(f"Data pregão(\n.*)", page_text)
                print(match)
                if match:
                    data1 = match.group(1).strip()
                    # Find the next empty row in the sheet
                    next_row = sheet.max_row + 1 # type: ignore
                    # Write the data to the sheet
                    sheet.cell(row=next_row, column=1, value=data1) # type: ignore
                    # print(data1)

                print_next = False

                for i, line in enumerate(page_text2):
                    if 'Total líquido da nota' in line:
                        print_next = True
                        valor1 = page_text2[i + 6]
                        valor1 = valor1.split(' ')
                        print(valor1)
                        print_next = False
                        valor1[0] = float(valor1[0].replace('.', '').replace(',', '.')) # type: ignore
                        if valor1[1] == 'D':
                            valor1[0] = valor1[0] * -1
                            print(valor1[0], valor1[1])
                            print_next = False
                            next_row = sheet.max_row
                            sheet.cell(row=next_row, column=2, value=valor1[0])
                            sheet.cell(row=next_row, column=3, value=valor1[1]) 
                        else:
                            print(valor1[0], valor1[1])
                            print_next = False
                            next_row = sheet.max_row
                            sheet.cell(row=next_row, column=2, value=valor1[0])
                            sheet.cell(row=next_row, column=3, value=valor1[1]) 
                     
                    elif 'Líquido para' in line:
                        print_next = True
                        valor2 = page_text2[i - 1]
                        valor3 = page_text2[i + 1]
                        print(valor2, valor3)
                        valor2 = float(valor2.replace('.', '').replace(',', '.'))
                        if valor3 == 'D':
                            valor2 = valor2 * -1
                            print(valor2, valor3)
                            print_next = False
                            next_row = sheet.max_row
                            sheet.cell(row=next_row, column=2, value=valor2)
                            sheet.cell(row=next_row, column=3, value=valor3)
                        else:
                            print_next = False
                            next_row = sheet.max_row
                            sheet.cell(row=next_row, column=2, value=valor2)
                            sheet.cell(row=next_row, column=3, value=valor3)     
# Save the Excel workbook
wb.save(r'BackTesting\notas\resumo.xlsx')

# Carregar o arquivo Excel
nome_arquivo = r'BackTesting\notas\resumo.xlsx'

df = pd.read_excel(nome_arquivo)

# Converter a coluna 'Data' para o tipo de data
df['Data'] = pd.to_datetime(df['Data'], dayfirst=True)

# Agrupar e somar os valores por mês e ano
df_agrupado = df.groupby(df['Data'].dt.to_period('M'))['Valor Total'].sum()
# Mudar o formato da data
df_agrupado.to_excel(r'BackTesting\notas\resumo_agrupado.xlsx')