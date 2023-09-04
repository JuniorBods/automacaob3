import os
from openpyxl import Workbook, load_workbook
import PyPDF2
import re
import pandas as pd

# Define the folder where the PDF files are located
pdf_folder = r'C:\Python\pythonProject\automacaob3\BackTesting\notastoro'

# Create a new Excel workbook
wb = Workbook()
sheet = wb.active

# Add headers to the columns
sheet['A1'] = 'Data'
sheet['B1'] = 'Valor Total'


# Iterate over all the files in the folder
for filename in os.listdir(pdf_folder):
    if filename.endswith('.pdf'):
        file_path = os.path.join(pdf_folder, filename)
        # Open the PDF file using PdfReader
        with open(file_path, 'rb') as pdf_file:
            # print(pdf_file.read())
            reader = PyPDF2.PdfReader(pdf_file)
            for page in reader.pages:
                # Extraia o texto da página atual
                page_text = page.extract_text()
                # print(page_text)
                match_data = re.search(r'\d{2}\/\d{2}\/\d{4}', page_text)
                if match_data:
                    # Obtém a data correspondente ao match
                    date = match_data.group()
                    # Find the next empty row in the sheet
                    next_row = sheet.max_row + 1
                    # Write the data to the sheet
                    sheet.cell(row=next_row, column=1, value=date)
                    # print(date)

                match_valor = re.search(
                    r'[0-9]{1,3}(?:[,]\d{3})*[.][0-9]{2} [CD] [CD] [0-9]{1,3}(?:[,]\d{3})*[.][0-9]{2} [0-9]{1,3}(?:[,]\d{3})*[.][0-9]{2}', page_text)
                # print(match_valor)
                if match_valor:
                    # Obtém a data correspondente ao match
                    valor = match_valor.group()
                    valor_DC = valor.split(' ')[-3]
                    valor_final = float(valor.replace(',', '').split(' ')[-1])
                    if valor_DC == 'D':
                        valor_final = valor_final * -1
                        next_row = sheet.max_row
                        sheet.cell(row=next_row, column=2, value=valor_final)
                        # print(valor_final)
                    else:
                        next_row = sheet.max_row
                        sheet.cell(row=next_row, column=2, value=valor_final)
                        # print(valor_final)

# Save the Excel workbook
wb.save(r'BackTesting\notastoro\resumo.xlsx')

# Carregar o arquivo Excel
nome_arquivo = r'BackTesting\notastoro\resumo.xlsx'

df = pd.read_excel(nome_arquivo)
print(df)
# Converter a coluna 'Data' para o tipo de data
df['Data'] = pd.to_datetime(df['Data'], dayfirst=True)

# Agrupar e somar os valores por mês e ano
df_agrupado = df.groupby(df['Data'].dt.to_period('M'))['Valor Total'].sum()
print(df_agrupado)
# Mudar o formato da data
df_agrupado.to_excel(r'BackTesting\notastoro\resumo_agrupado.xlsx')
